#!/usr/bin/env python3
"""Apply per-module hospital links for CUHK & SZUFH from pricedata / source xlsx."""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "db.js"
SURGERY_CSV = ROOT / "pricedata" / "手术费用_CUHKMC-整理后表格.csv"
OUTPATIENT_CSV = ROOT / "pricedata" / "门诊费_CUHKMC-门诊费.csv"
SURGERY_XLSX = Path("/Users/boyiyang/Downloads/手术费用_CUHKMC.xlsx")
OUTPATIENT_XLSX = Path("/Users/boyiyang/Downloads/门诊费_CUHKMC.xlsx")
LINKS_JSON = Path(__file__).resolve().parent / "hospital-links.json"

HOSPITAL_IDS = {
    "香港中文大學醫院": "cuhk",
    "深圳新風和睦家醫院": "szufh",
}

HOSPITAL_ORDER = [
    "szufh", "cuhk", "hksh", "ghk", "matilda", "sth", "baptist", "union",
    "canossa", "sph", "pbh", "evangel", "twah", "hkah",
]

FLAT_MODULES = {"outpatient", "outpatientSpecialty", "ward"}


def is_hospital_id(name: str) -> bool:
    return name in HOSPITAL_ORDER

# SZUFH: pricedata 链接列为「查看頁面」时，按科室映射至官网套餐页
SZUFH_DEPT_LINKS = {
    "內視鏡中心": "https://www.szufh.hk/xiaohuaneijing.html",
    "婦產科": "https://www.szufh.hk/shoushusf.html",
    "泌尿外科": "https://www.szufh.hk/shoushusf.html",
    "普通外科": "https://www.szufh.hk/shoushusf.html",
    "骨科": "https://www.szufh.hk/shoushusf.html",
    "麻醉科": "https://www.szufh.hk/shoushusf.html",
    "耳鼻喉科": "https://www.szufh.hk/shoushusf.html",
    "眼科": "https://www.szufh.hk/shoushusf.html",
    "整形外科": "https://www.szufh.hk/shoushusf.html",
    "心血管內科": "https://www.szufh.hk/shoushusf.html",
}

CUHK_ENDOSCOPY_PKG = "https://www.cuhkmc.hk/sc/medical-packages/cumc-medical-package/endoscopy-package-fees"
CUHK_REFERENCE = "https://www.cuhkmc.hk/sc/fees-and-charges/price-transparency/reference-charges-for-common-surgical-procedures"
CUHK_DEPT_FALLBACKS = {
    "婦產科": "https://www.cuhkmc.hk/sc/fees-and-charges/maternity",
    "骨科": "https://www.cuhkmc.hk/sc/medical-packages/cumc-medical-package/orthopaedics",
    "普通外科": "https://www.cuhkmc.hk/sc/medical-packages/cumc-medical-package/general-surgery",
    "眼科": CUHK_REFERENCE,
    "耳鼻喉科": CUHK_REFERENCE,
}


def cuhk_dept_fallback(row: dict) -> str | None:
    dept = row["dept"]
    pkg = row["package"]
    if dept == "內視鏡中心":
        if any(k in pkg for k in ("鎮靜麻醉", "監測麻醉", "鎮靜/監測")):
            return CUHK_ENDOSCOPY_PKG
        return CUHK_REFERENCE
    return CUHK_DEPT_FALLBACKS.get(dept)

OUTPATIENT_LINKS = {
    "outpatient.cuhk": "https://www.cuhkmc.hk/sc/fees-and-charges/emergency-medicine-centre",
    "outpatientSpecialty.cuhk": "https://www.cuhkmc.hk/sc/fees-and-charges/specialist-outpatient-clinic",
    "outpatient.szufh": "https://www.szufh.hk/fuwusf.html",
    "outpatientSpecialty.szufh": "https://www.szufh.hk/yiliaoquan.html",
}

# Mirrors scripts/build-surgery-modules.js ROW_MAP (match, module, procedure)
ROW_MAP = [
    {"match": lambda r: r["hospital"] == "cuhk" and "支氣管鏡" in r["package"] and "日間" in r["package"], "module": "imaging", "procedure": "bronchoscopy"},
    {"match": lambda r: r["hospital"] == "cuhk" and "胃鏡檢查 (鎮靜麻醉)" in r["package"], "module": "imaging", "procedure": "gastroscopy"},
    {"match": lambda r: r["hospital"] == "szufh" and "無痛胃鏡" in r["package"], "module": "imaging", "procedure": "gastroscopy"},
    {"match": lambda r: r["hospital"] == "cuhk" and "結腸內視鏡檢查 (鎮靜麻醉)" in r["package"], "module": "imaging", "procedure": "colonoscopy"},
    {"match": lambda r: r["hospital"] == "szufh" and "無痛腸鏡" in r["package"], "module": "imaging", "procedure": "colonoscopy"},
    {"match": lambda r: r["hospital"] == "cuhk" and "胃鏡及大腸鏡" in r["package"] and "日間" in r["package"], "module": "imaging", "procedure": "dual_scope"},
    {"match": lambda r: r["hospital"] == "szufh" and "雙鏡聯合" in r["package"], "module": "imaging", "procedure": "dual_scope"},
    {"match": lambda r: "子宮頸病變" in r["package"], "module": "gynecology", "procedure": "cervical_treatment"},
    {"match": lambda r: "子宮鏡診治" in r["package"], "module": "gynecology", "procedure": "hysteroscopy"},
    {"match": lambda r: "子宮肌瘤切除" in r["package"], "module": "gynecology", "procedure": "myomectomy"},
    {"match": lambda r: "子宮切除術" in r["package"], "module": "gynecology", "procedure": "hysterectomy"},
    {"match": lambda r: "輸卵管及宮外孕" in r["package"], "module": "gynecology", "procedure": "tubal_ectopic"},
    {"match": lambda r: "卵巢囊腫切除" in r["package"], "module": "gynecology", "procedure": "ovarian_cyst"},
    {"match": lambda r: "避孕及終止妊娠" in r["package"], "module": "gynecology", "procedure": "contraception"},
    {"match": lambda r: "自然分娩" in r["package"], "module": "gynecology", "procedure": "normal_delivery"},
    {"match": lambda r: "剖腹分娩" in r["package"], "module": "gynecology", "procedure": "c_section"},
    {"match": lambda r: "膽囊切除" in r["package"], "module": "generalSurgery", "procedure": "cholecystectomy"},
    {"match": lambda r: r["hospital"] == "cuhk" and "乳房腫塊切除" in r["package"], "module": "generalSurgery", "procedure": "breast_lump"},
    {"match": lambda r: r["hospital"] == "cuhk" and "包皮環切" in r["package"], "module": "generalSurgery", "procedure": "circumcision"},
    {"match": lambda r: r["hospital"] == "cuhk" and "甲狀腺細針穿刺" in r["package"], "module": "generalSurgery", "procedure": "thyroid_fna"},
    {"match": lambda r: "偏側甲狀腺切除" in r["package"], "module": "generalSurgery", "procedure": "hemithyroidectomy"},
    {"match": lambda r: "局部/次全/全甲狀腺" in r["package"], "module": "generalSurgery", "procedure": "thyroidectomy"},
    {"match": lambda r: "腹腔疝氣修補" in r["package"], "module": "generalSurgery", "procedure": "hernia_abdominal"},
    {"match": lambda r: "單側腹股溝疝氣" in r["package"], "module": "generalSurgery", "procedure": "hernia_unilateral"},
    {"match": lambda r: "雙側腹股溝疝氣" in r["package"], "module": "generalSurgery", "procedure": "hernia_bilateral"},
    {"match": lambda r: "闌尾切除" in r["package"], "module": "generalSurgery", "procedure": "appendectomy"},
    {"match": lambda r: "痔瘡" in r["package"], "module": "generalSurgery", "procedure": "hemorrhoid"},
    {"match": lambda r: r["hospital"] == "szufh" and "甲狀腺處置" in r["package"], "module": "generalSurgery", "procedure": "thyroid_surgery"},
    {"match": lambda r: "輸液港" in r["package"], "module": "generalSurgery", "procedure": "port_a_cath"},
    {"match": lambda r: "乳腺抽針" in r["package"], "module": "generalSurgery", "procedure": "breast_biopsy"},
    {"match": lambda r: "乳腺外科手術" in r["package"], "module": "generalSurgery", "procedure": "breast_surgery"},
    {"match": lambda r: "關節鏡手術 (肩部" in r["package"], "module": "orthopedics", "procedure": "shoulder_arthroscopy"},
    {"match": lambda r: "全肩關節置換" in r["package"], "module": "orthopedics", "procedure": "shoulder_replacement"},
    {"match": lambda r: "鎖骨/手腕橈骨" in r["package"], "module": "orthopedics", "procedure": "orif_upper_limb"},
    {"match": lambda r: "腕管解除" in r["package"], "module": "orthopedics", "procedure": "carpal_tunnel"},
    {"match": lambda r: "板機狀指" in r["package"], "module": "orthopedics", "procedure": "trigger_finger"},
    {"match": lambda r: "關節鏡手術 (膝部" in r["package"], "module": "orthopedics", "procedure": "knee_arthroscopy"},
    {"match": lambda r: "全人工膝關節置換" in r["package"], "module": "orthopedics", "procedure": "knee_replacement"},
    {"match": lambda r: "髕骨/單雙踝" in r["package"], "module": "orthopedics", "procedure": "orif_lower_limb"},
    {"match": lambda r: "髖關節全關節置換" in r["package"], "module": "orthopedics", "procedure": "hip_replacement"},
    {"match": lambda r: "跟腱修補" in r["package"], "module": "orthopedics", "procedure": "achilles_ankle"},
    {"match": lambda r: "人工關節置換術 (全膝" in r["package"], "module": "orthopedics", "procedure": "joint_replacement"},
    {"match": lambda r: "脊柱手術" in r["package"], "module": "orthopedics", "procedure": "spine_surgery"},
    {"match": lambda r: "骨科小手術" in r["package"], "module": "orthopedics", "procedure": "sports_ortho"},
    {"match": lambda r: "微型喉鏡" in r["package"], "module": "ent", "procedure": "micro_laryngoscopy"},
    {"match": lambda r: r["hospital"] == "cuhk" and "扁桃體切除" in r["package"], "module": "ent", "procedure": "tonsillectomy"},
    {"match": lambda r: "腺樣體" in r["package"], "module": "ent", "procedure": "adenoid_tonsil"},
    {"match": lambda r: "鼻竇炎及鼻中隔" in r["package"], "module": "ent", "procedure": "sinus_surgery"},
    {"match": lambda r: "鼓膜修補" in r["package"], "module": "ent", "procedure": "tympanoplasty"},
    {"match": lambda r: r["hospital"] == "cuhk" and "白內障" in r["package"], "module": "ophthalmology", "procedure": "cataract"},
    {"match": lambda r: r["hospital"] == "szufh" and "白內障" in r["package"], "module": "ophthalmology", "procedure": "cataract"},
    {"match": lambda r: "斜視手術" in r["package"], "module": "ophthalmology", "procedure": "strabismus"},
    {"match": lambda r: "前列腺診療" in r["package"], "module": "urology", "procedure": "prostate"},
    {"match": lambda r: "結石碎石" in r["package"], "module": "urology", "procedure": "kidney_stone"},
    {"match": lambda r: "尿動力檢查" in r["package"], "module": "urology", "procedure": "urodynamics"},
    {"match": lambda r: "男科處置" in r["package"], "module": "urology", "procedure": "andrology"},
    {"match": lambda r: "疼痛管理" in r["package"], "module": "painManagement", "procedure": "nerve_block"},
    {"match": lambda r: "脊柱內鏡診療" in r["package"], "module": "painManagement", "procedure": "spine_endoscopy"},
    {"match": lambda r: "脊髓電刺激" in r["package"], "module": "painManagement", "procedure": "scs_implant"},
    {"match": lambda r: "冠狀動脈造影" in r["package"], "module": "cardiology", "procedure": "pci"},
    {"match": lambda r: "美容縫合" in r["package"], "module": "plastics", "procedure": "laceration_repair"},
]


def norm_pkg(text: str) -> str:
    return re.sub(r"\s+", "", str(text or ""))


def parse_surgery_csv(text: str) -> list[dict]:
    rows = []
    for line in text.strip().split("\n")[1:]:
        parts = []
        cur = ""
        in_quote = False
        for ch in line:
            if ch == '"':
                in_quote = not in_quote
                continue
            if ch == "," and not in_quote:
                parts.append(cur)
                cur = ""
                continue
            cur += ch
        parts.append(cur)
        if len(parts) < 8:
            continue
        name = parts[0]
        rows.append({
            "hospitalName": name,
            "hospital": HOSPITAL_IDS.get(name),
            "dept": parts[1],
            "package": parts[2],
            "linkRaw": parts[7],
        })
    return [r for r in rows if r["hospital"]]


def load_surgery_xlsx_links() -> dict[tuple, str]:
    if not SURGERY_XLSX.exists():
        return {}
    wb = load_workbook(SURGERY_XLSX, data_only=True)
    ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        h = ws.cell(r, 1).value
        dept = ws.cell(r, 2).value
        pkg = ws.cell(r, 3).value
        link = ws.cell(r, 8).value
        if link and str(link).startswith("http"):
            out[(h, dept, norm_pkg(pkg))] = str(link).strip()
    return out


def resolve_surgery_link(row: dict, xlsx_links: dict[tuple, str]) -> str | None:
    hname = row["hospitalName"]
    dept = row["dept"]
    pkg = row["package"]

    if row["hospital"] == "cuhk":
        fallback = cuhk_dept_fallback(row)
        key = (hname, dept, norm_pkg(pkg))
        if key in xlsx_links:
            return xlsx_links[key]
        for (xh, xd, xp), link in xlsx_links.items():
            if xh != hname or xd != dept:
                continue
            np, nx = norm_pkg(pkg), xp
            if np in nx or nx in np or np[:8] == nx[:8]:
                return link
        return fallback

    if row["hospital"] == "szufh":
        return SZUFH_DEPT_LINKS.get(dept, "https://www.szufh.hk/shoushusf.html")
    return None


def build_surgery_link_map(rows: list[dict]) -> dict[str, dict[str, dict[str, str]]]:
    xlsx_links = load_surgery_xlsx_links()
    store: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

    for row in rows:
        link = resolve_surgery_link(row, xlsx_links)
        if not link:
            continue
        for rule in ROW_MAP:
            if not rule["match"](row):
                continue
            mod, proc = rule["module"], rule["procedure"]
            hid = row["hospital"]
            store[mod][proc][hid] = link
    return store


def update_surgery_csv(rows: list[dict], xlsx_links: dict[tuple, str]) -> None:
    with open(SURGERY_CSV, encoding="utf-8") as f:
        text = f.read()
    lines = text.strip().split("\n")
    header = lines[0]
    out_lines = [header]
    idx = 0
    for line in lines[1:]:
        if idx >= len(rows):
            out_lines.append(line)
            continue
        row = rows[idx]
        link = resolve_surgery_link(row, xlsx_links)
        idx += 1
        if not link:
            out_lines.append(line)
            continue
        # replace last field
        parts = []
        cur = ""
        in_quote = False
        for ch in line:
            if ch == '"':
                in_quote = not in_quote
                cur += ch
                continue
            if ch == "," and not in_quote:
                parts.append(cur)
                cur = ""
                continue
            cur += ch
        parts.append(cur)
        parts[-1] = link
        rebuilt = []
        for i, p in enumerate(parts):
            if "," in p or "\n" in p:
                rebuilt.append('"' + p.replace('"', '""') + '"')
            else:
                rebuilt.append(p)
        out_lines.append(",".join(rebuilt))
    SURGERY_CSV.write_text("\n".join(out_lines) + "\n", encoding="utf-8")


def update_outpatient_csv() -> None:
    mapping = {
        ("香港中文大學醫院", "普通門診"): OUTPATIENT_LINKS["outpatient.cuhk"],
        ("香港中文大學醫院", "專科門診"): OUTPATIENT_LINKS["outpatientSpecialty.cuhk"],
        ("深圳新風和睦家醫院", "普通門診"): OUTPATIENT_LINKS["outpatient.szufh"],
        ("深圳新風和睦家醫院", "專科門診"): OUTPATIENT_LINKS["outpatientSpecialty.szufh"],
    }
    with open(OUTPATIENT_CSV, encoding="utf-8") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    link_col = len(header) - 1
    for row in rows[1:]:
        if len(row) <= link_col:
            continue
        key = (row[0], row[1])
        if key in mapping:
            row[link_col] = mapping[key]
    with open(OUTPATIENT_CSV, "w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(rows)


def flatten_link_map(surgery_map: dict) -> dict[str, str]:
    flat: dict[str, str] = dict(OUTPATIENT_LINKS)
    for mod, procs in surgery_map.items():
        for proc, hospitals in procs.items():
            for hid, link in hospitals.items():
                flat[f"{mod}.{proc}.{hid}"] = link
    return flat


def patch_db_js(flat_links: dict[str, str]) -> int:
    lines = DB_PATH.read_text(encoding="utf-8").split("\n")
    out = []
    in_modules = False
    module = None
    procedure = None
    in_target_hospital = False
    hospital_indent = ""
    patched = 0

    for line in lines:
        if line.strip() == "modules: {":
            in_modules = True
            out.append(line)
            continue
        if in_modules and line.strip().startswith("searchIndex:"):
            in_modules = False
            module = None
            procedure = None

        if not in_modules:
            out.append(line)
            continue

        if in_target_hospital:
            if re.match(r"^\s+link: ", line):
                continue
            if line.rstrip().endswith("},") or line.rstrip() == "}":
                if hospital_indent and line.startswith(hospital_indent):
                    in_target_hospital = False
                    hospital_indent = ""
            out.append(line)
            continue

        m4 = re.match(r"^    (\w+): \{$", line)
        if m4:
            module = m4.group(1)
            procedure = None
            out.append(line)
            continue

        m6 = re.match(r"^      (\w+): \{$", line)
        if m6 and module:
            name = m6.group(1)
            if is_hospital_id(name):
                if name not in ("cuhk", "szufh"):
                    out.append(line)
                    continue
                indent, hid = "      ", name
                key = f"{module}.{hid}" if module in FLAT_MODULES else (
                    f"{module}.{procedure}.{hid}" if procedure else None
                )
                link = flat_links.get(key) if key else None
                out.append(line)
                if link:
                    out.append(f'{indent}  link: "{link}",')
                    in_target_hospital = True
                    hospital_indent = indent
                    patched += 1
                continue
            if module not in FLAT_MODULES:
                procedure = name
            out.append(line)
            continue

        out.append(line)

    DB_PATH.write_text("\n".join(out), encoding="utf-8")
    return patched


def main() -> None:
    surgery_rows = parse_surgery_csv(SURGERY_CSV.read_text(encoding="utf-8"))
    xlsx_links = load_surgery_xlsx_links()
    update_surgery_csv(surgery_rows, xlsx_links)
    update_outpatient_csv()
    surgery_rows = parse_surgery_csv(SURGERY_CSV.read_text(encoding="utf-8"))
    surgery_map = build_surgery_link_map(surgery_rows)
    flat = flatten_link_map(surgery_map)
    LINKS_JSON.write_text(json.dumps({"surgery": surgery_map, "flat": flat}, indent=2, ensure_ascii=False), encoding="utf-8")
    count = patch_db_js(flat)
    print(f"Patched {count} hospital link entries in db.js")
    print(f"Wrote {LINKS_JSON}")


if __name__ == "__main__":
    main()
